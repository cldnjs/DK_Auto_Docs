from openpyxl import *


def load_excel(filename, sheet_name, read_only=False, data_only=False):
    """
    엑셀 파일을 로드
    :param filename: String
    :param sheet_name: String
    :param read_only: Bool
    :param data_only: Bool
    :return: WorkBook, WorkSheet
    """
    wb = load_workbook(filename, read_only, data_only)
    ws = wb[sheet_name]

    return wb, ws


def load_column_data(sheet, col_num, start_num):
    """
    지정한 행의 지정한 시작지점부터의 데이터들을 반환
    :param sheet:
    :param col_num: Int
    :param start_num: Int
    :return: List
    """
    column_data = []
    for row in range(start_num, sheet.max_row+1):
        data = sheet.cell(row=row, column=col_num).value
        if data is None:
            pass
        else:
            column_data.append(data)

    return column_data


def load_row_data(sheet, row_num, start_num):
    """
    지정한 열의 지정된 시작지점부터의 데이터들을 반환
    :param sheet: WorkSheet
    :param row_num: Int
    :param start_num: Int
    :return: List
    """
    row_data = []
    for col in range(start_num, sheet.max_column+1):
        data = sheet.cell(row=row_num, column=col).value
        if data is None:
            pass
        else:
            row_data.append(data)

    return row_data
